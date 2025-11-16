
import click
import json
import csv
from pathlib import Path

from datetime import datetime
from loguru import logger

import openpyxl
from openpyxl.utils import get_column_letter

from collections import defaultdict

from watools.core.api import get_events, get_event_details, get_event_registrants, \
    get_default_membership_levels, get_default_membergroups, \
    get_contacts, register_contacts_to_event

from watools.core.utils import filter_events, list_events, list_event_details


@click.group("events", invoke_without_command=True)
@click.option('--event-id', type=int, help='Filter by specific event ID')
@click.option('--all', 'show_all', is_flag=True, help='Show all events, bypass default date filter')
@click.option('--future', is_flag=True, help='Show only future events')
@click.option('--year', type=int, help='Filter by event start year')
@click.option('--month', type=int, help='Filter by event start month')
@click.option('--after', type=click.DateTime(), help='Only events after this date')
@click.option('--before', type=click.DateTime(), help='Only events before this date')
@click.option('--query', type=str, help='Ad hoc query expression (e.g., \'ConfirmedRegistrationsCount > 5 and "Diamond" in Name\')')
@click.option('--as-json', is_flag=True, default=False, help='Output events as JSON')
@click.pass_context
def cmd(ctx, event_id, show_all,future, year, month, after, before, query, as_json):
    """Manage Wild Apricot events."""

    ctx.ensure_object(dict)
    account_id = ctx.obj.get("account_id")
    if not account_id:
        logger.error("No account ID provided. Use --account-id or configure it.")
        return

    logger.debug(f"Fetching events for account {account_id}")
    event_data = get_events(account_id)
    events = event_data.get("Events", [])

    # Filter by ID directly
    if event_id:

        events = [e for e in events if str(e.get("Id")) == str(event_id)]

        if not events:
            click.echo(f"No event found with ID: {event_id}")
            return

        if ctx.invoked_subcommand:
            ctx.obj["event_id"] = str(event_id)
            logger.debug(f"Leaving group commands to run subcommand: {ctx.invoked_subcommand}")
            return

        # Show single event details
        event = get_event_details(event_id, account_id=account_id)
        logger.debug( event )
        if event:
            if as_json:
                click.echo(json.dumps(event, indent=2))
            else:
                list_event_details(event)
        else:
            click.echo(f"Failed to load details for event ID {event_id}")
        return

    # Apply filters
    events = filter_events(
        events,
        show_all=show_all,
        future=future,
        year=year,
        month=month,
        after=after,
        before=before,
        query=query
    )

    if not events:
        click.echo("No events found.")
        return

    if as_json:
        click.echo(json.dumps(events, indent=2))
    else:
        click.echo("")  # spacing
        list_events(
            events,
            columns=[
                "Id",
                {"wat_start_date": "Date"},
                {"wat_start_day": "Day"},
                {"wat_start_time": "Time"},
                {"wat_confirmed_and_limit":"Cnf/Max"},
                {"Name": "Title"},
            ]
        )    


VALID_STATUSES = ['Active', 'Lapsed', 'PendingRenewal', 'PendingNew']

def parse_statuses(ctx, param, values):
    """Flatten multiple values and comma-separated values, validate each one."""
    result = []
    for value in values:
        parts = [v.strip() for v in value.split(',')]
        for part in parts:
            if part not in VALID_STATUSES:
                raise click.BadParameter(f"Invalid status: '{part}'. Valid options: {', '.join(VALID_STATUSES)}")
            result.append(part)
    return result


@cmd.command()
@click.option('--event-id', type=int, required=False, default=None, help='Event ID for auto-registration')
@click.option(
    '--use-status',
    callback=parse_statuses,
    multiple=True,
    default=('Active',),
    help='Filter auto-register contact IDs by status: Active, Lapsed, PendingRenewal, or PendingNew.'
)
@click.option('--confirm', is_flag=True, default=False, help='Confirm action before proceeding')
@click.pass_context
def auto_register(ctx, event_id, use_status, confirm):
    """ Autoregister contacts to event """

    account_id = ctx.obj.get('account_id')
    if not account_id:
        logger.error("No account ID provided. Use --account-id or add to config.toml.")
        return

    event_id = ctx.obj.get('event_id')
    if not event_id:
        logger.error("No event_id provided. Use --event-id to specify an event.")
        return

    try:
        # Fetch event details (dict), look for AccessControl
        event = get_event_details( event_id, account_id=account_id )
        if not event:
            logger.error(f"No event found with ID: {event_id}")
            return
        logger.info(f"---- Auto-registration for: {event["Name"]} ----")

        # Fetch registrants to current event (list of dict)
        registrants = get_event_registrants( event_id, account_id=account_id )


        # Fetch all contacts/members across entire account.
        contacts = get_contacts( account_id )
        if not contacts:
            click.echo(f"No contacts found for account ID {account_id}.")
            return
        logger.debug( f"Contacts: {len(contacts)}" )

        # Fetch default membership levels IDs.  These are for the entire account, not just event.
        default_membership_levels = get_default_membership_levels( account_id )
        default_membership_levels = {item["Id"]: item for item in default_membership_levels}
        if not default_membership_levels:
            click.echo(f"No membership levels found for account ID {account_id}.")
            return
        logger.debug( default_membership_levels)
        default_membership_level_ids = default_membership_levels.keys()
        logger.debug( f"Membership level ids: {default_membership_level_ids}" )

        # Fetch member groups ids.  These are for entire account, not just event.
        default_membergroups = get_default_membergroups( account_id )
        default_membergroups = {item["Id"]: item for item in default_membergroups}
        if not default_membergroups:
            click.echo(f"No member groups found for account ID {account_id}.")
            return
        logger.debug( default_membergroups )
        default_membergroup_ids = default_membergroups.keys()
        logger.debug( f"Member group ids: {default_membergroup_ids}" )

        # Event Access control: by membership level and member group
        # This determines who can see the event.  If you can see it, you can be auto-registered for it!
        access_control = event.get("Details",{}).get("AccessControl")
        if not access_control:
            logger.error(f"No access control found for event ID {event_id}.")
            return 
        
        logger.debug(f"Access control: {json.dumps( access_control,indent=2)}")

        # Potential pools of registrants: Everyone unless restricted by access control.  Member level AND groups.
        membership_levels_ids = default_membership_level_ids
        if not access_control["AvailableForAnyLevel"]:
            membership_levels_ids = [item["Id"] for item in access_control.get("AvailableForLevels",[])]
        logger.debug(f"Membership levels ids: {membership_levels_ids}")
        for id in membership_levels_ids:
            logger.info(f"Autoregister level: {default_membership_levels.get(id,{}).get("Name")}")

        membergroup_ids = default_membergroup_ids
        if not access_control["AvailableForAnyGroup"]:
            membergroup_ids = [item["Id"] for item in access_control.get("AvailableForGroups",[])]
        logger.debug(f"Member group ids: {membergroup_ids}")
        for id in membergroup_ids:
            logger.info(f"Autoregister group: {default_membergroups.get(id,{}).get("Name")} ({default_membergroups.get(id,{}).get("ContactsCount")})")
    


        members_ids_by_level = [ contact["Id"] for contact in contacts if contact["MembershipLevelId"] in membership_levels_ids ]
        logger.debug(f"Count of members ids by level: {len(members_ids_by_level)}")

        #logger.debug(f"{json.dumps([contact for contact in contacts if "Leonard" in contact["DisplayName"]],indent=2)}")

        def contact_in_group(contact, group_ids):
            for field in contact.get("FieldValues", []):
                if field.get("SystemCode") == "Groups":
                    for group in field.get("Value", []):
                        if group.get("Id") in group_ids:
                            return True
            return False
        
        member_ids_by_group = [c["Id"] for c in contacts if contact_in_group(c, membergroup_ids)]
        logger.debug(f"Count of member_ids_by_group: {len(member_ids_by_group)}")

        current_registrant_ids = [c.get("Contact",{}).get("Id") for c in registrants]

        # build list of potential registrants by combining level and group ids and removing current registrants
        temp_ids = list(set(members_ids_by_level + member_ids_by_group))
        potential_registrant_ids = [cid for cid in temp_ids if cid not in current_registrant_ids]
        
        logger.info(f"---- Counts of current and potential registrants ----")
        logger.info(f"Count of current registrants: {len(current_registrant_ids)}")
        logger.info(f"Count of potential registrants: {len(potential_registrant_ids)}")

        # Given list potential_registrant_ids, create sublists by membership status (e.g., Active, etc.)
        status_groups = defaultdict(list)
        for contact in contacts:
            cid = contact.get("Id")
            if cid in potential_registrant_ids:
                status = contact.get("Status","Unknown")
                status_groups[status].append(cid)

        logger.info("---- Breakdown of potential registrants ----")
        for key in status_groups.keys():
            logger.info(f"Count: {len(status_groups[key]):5} - {key}")         
        logger.info(f"Using status(s): {use_status}")  

        # Derive list of potential ticket types

        registration_types = event.get("Details",{}).get("RegistrationTypes",{})
        if not registration_types:
            click.echo(f"No registration types found for event ID {event_id}.")
            return
        logger.debug(f"Registration types: {json.dumps(registration_types,indent=2)}")

        registration_type_ids = [item["Id"] for item in registration_types if "auto-register" in item["Name"].lower() ]
        registration_type_names = [item["Name"] for item in registration_types if "auto-register" in item["Name"].lower() ]
        logger.debug( f"Registration type IDs with 'auto-register' in name: { registration_type_ids }" )

        if not registration_type_ids:
            logger.error(f"No auto-register registration types found for event ID: {event_id}.")
            logger.error(f"Review event and add one registration type with 'auto-register' in the name.")
            return
        
        if len(registration_type_ids) > 1:
            logger.error(f"Multiple auto-register registration types found for event ID: {event_id}.")
            logger.error(f"Review event and delete extra auto-register events.")
            return

        logger.info(f"Assigning registrants to: {registration_type_names[0]}")

        #logger.debug(f"Contact: {json.dumps([contact for contact in contacts if contact['DisplayName'] in ["Leonard, John"]], indent=2)}")

        if confirm:
            if status_groups:
                for key in use_status:
                    if len(status_groups[key]) > 0:


                        result = register_contacts_to_event(
                            contact_ids=status_groups[key],      # list of contact ids from specific status group
                            event_id=event_id,                   # current event_id
                            account_id=account_id,               # current account_id
                            reg_type_id=registration_type_ids[0] # auto-register registration type id
                        )

                        click.echo(f"[{key}] registrations: {len(result["success"])} successful.")
                        click.echo(f"[{key}] registrations: {len(result["failed"])} failed.")
                    else:
                        logger.debug(f"Key {key} has no records to process.")
            else:
                click.echo("No records to process.")
        else:
            logger.info("No action taken.  use --confirm to process")
        return
        event_details = get_event_details( account_id, event_id )
        access_control = event_details.get("Details",{}).get("AccessControl",{})
        logger.debug( json.dumps(access_control,indent=2) )


        if as_json:
            click.echo(json.dumps(registrants, indent=2))
        else:
            for reg in registrants:
                name = reg.get("DisplayName", "Unknown")
                click.echo(f"Name: {name}")

    
    except Exception as e:
        #logger.error(f"Error fetching registrants: {e}")
        click.echo(f"Error: {e}")


@cmd.command("dump-registrations")
@click.option("--event-id", type=int,
              help="Event ID to dump registrations for (overrides outer --event-id).")
@click.option("--outfile", type=click.Path(dir_okay=False),
              required=True,
              help="XLSX filename to write (e.g., registrations.xlsx).")
@click.pass_context
def dump_registrations(ctx, event_id, outfile):
    """
    Dump event registrations to XLSX, including:
      • Guest count
      • Invoice totals (RegistrationFee / PaidSum / Balance)
      • Registration type name + ID
      • Contact flattening
      • Event metadata sheet
      • Styled headers and auto-width with max 50 chars
    """

    account_id = ctx.obj.get("account_id")
    if not account_id:
        logger.error("No account ID provided.")
        return

    # -------------------------------
    # Resolve Event ID
    # -------------------------------
    event_id = event_id or ctx.obj.get("event_id")
    if not event_id:
        logger.error("No event_id provided. Use --event-id.")
        return

    # -------------------------------
    # Event metadata
    # -------------------------------
    event = get_event_details(event_id, account_id=account_id)
    if not event:
        click.echo(f"No event found with ID {event_id}")
        return

    # -------------------------------
    # Registrations
    # -------------------------------
    regs_raw = get_event_registrants(event_id, account_id=account_id)
    regs = regs_raw.get("EventRegistrations", regs_raw) if isinstance(regs_raw, dict) else regs_raw

    logger.info(f"Found {len(regs)} registrations")

    # -------------------------------
    # Contacts lookup
    # -------------------------------
    contacts = get_contacts(account_id)
    contacts_by_id = {c["Id"]: c for c in contacts}

    # -------------------------------
    # Build rows
    # -------------------------------
    rows = []
    all_keys = set()

    for reg in regs:
        row = {}

        row["RegistrationId"] = reg.get("Id")
        row["EventId"] = event_id

        # --- Registration type ---
        rt = reg.get("RegistrationType", {})
        row["RegistrationTypeId"] = rt.get("Id") or reg.get("RegistrationTypeId")
        row["RegistrationTypeName"] = rt.get("Name")

        # --- Registration status ---
        row["Status"] = reg.get("Status")
        row["IsCheckedIn"] = reg.get("IsCheckedIn")
        row["RegistrationDate"] = reg.get("RegistrationDate")

        # --- Correct Guest Count ---
        guest_info = reg.get("GuestRegistrationsSummary", {})
        row["GuestCount"] = guest_info.get("NumberOfGuests", 0)
        row["TicketCount"] = row["GuestCount"] + 1

        # --- Correct Invoice Totals ---
        fee = reg.get("RegistrationFee") or 0
        paid = reg.get("PaidSum") or 0

        row["InvoiceTotal"] = fee
        row["TotalPaid"] = paid
        row["InvoiceBalance"] = max(fee - paid, 0)

        # --- Contact info ---
        contact_id = reg.get("Contact", {}).get("Id")
        row["ContactId"] = contact_id

        # -------------------------------
        # Registration FieldValues (address, city, state, zip, custom fields)
        # -------------------------------
        for field in reg.get("RegistrationFields", []):
            name = field.get("FieldName")
            value = field.get("Value")

            # Normalize various WA value formats
            if isinstance(value, dict):
                # { "Id": 22444671, "Label": "VA" }
                val = value.get("Label") or value.get("Value") or str(value)
            elif isinstance(value, list):
                # Multi-select fields: [{"Id":..., "Label":"Word of Mouth"}]
                val = ", ".join(
                    f.get("Label") or f.get("Value") or str(f)
                    for f in value
                )
            else:
                val = value

            # Create safe column name
            safe_name = (
                "Reg_" +
                name.strip()
                    .replace(" ", "_")
                    .replace("/", "_")
                    .replace("(", "")
                    .replace(")", "")
                    .replace("-", "_")
            )

            row[safe_name] = val
            all_keys.add(safe_name)


        contact = contacts_by_id.get(contact_id, {})
        row["DisplayName"] = contact.get("DisplayName")
        row["Email"] = contact.get("Email")
        row["MembershipLevel"] = contact.get("MembershipLevelName")
        row["ContactStatus"] = contact.get("Status")

        # Flatten simple contact fields
        for k, v in contact.items():
            if isinstance(v, (str, int, float, bool)) or v is None:
                row[f"Contact_{k}"] = v

        rows.append(row)
        all_keys.update(row.keys())

    columns = sorted(all_keys)

    # -------------------------------
    # Write XLSX
    # -------------------------------
    outfile = Path(outfile)
    outfile.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "registrations"

    header_font = openpyxl.styles.Font(bold=True)
    header_align = openpyxl.styles.Alignment(horizontal="center")

    # Header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_align

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col_name))

    # Auto-width with max 50 chars
    for idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for r in rows:
            val = r.get(col_name)
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

    # -------------------------------
    # Metadata Sheet
    # -------------------------------
    meta = wb.create_sheet("event_metadata")
    meta_fields = [
        ("Event ID", event.get("Id")),
        ("Title", event.get("Name")),
        ("StartDate", event.get("StartDate")),
        ("EndDate", event.get("EndDate")),
        ("Location", event.get("Location")),
        ("ConfirmedRegistrations", event.get("ConfirmedRegistrationsCount")),
        ("RegistrationLimit", event.get("RegistrationsLimit")),
        ("AccessControl", json.dumps(event.get("Details", {}).get("AccessControl", {}), indent=2)),
    ]

    for i, (k, v) in enumerate(meta_fields, start=1):
        meta.cell(row=i, column=1, value=k)
        meta.cell(row=i, column=2, value=v)

    wb.save(outfile)
    click.echo(f"✓ XLSX written: {outfile}")
